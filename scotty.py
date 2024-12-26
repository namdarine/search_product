import requests
import json
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from function import get_API_key, find_file
import os
import subprocess


key_file = './scotty.txt'
api_key_instance = get_API_key(key_file, 5)
api_key = api_key_instance.get_api_key(5).strip()

current_date = datetime.now().strftime("%Y-%m-%d")
last_update_date = (datetime.now() - timedelta(days=24)).strftime("%Y-%m-%dT%H:%M:%S.000Z")

url = "https://svcs.ebay.com/services/search/FindingService/v1"

headers = {
    "X-EBAY-SOA-OPERATION-NAME": "findItemsByKeywords",
    "X-EBAY-SOA-SERVICE-VERSION": "1.0.0",
    "X-EBAY-SOA-SECURITY-APPNAME": api_key,
    "X-EBAY-SOA-RESPONSE-DATA-FORMAT": "JSON"
}

all_items = []
page_number = 1

while True:
    params = {
            "keywords": "scotty cameron",
            "paginationInput.entriesPerPage": "100",
            "paginationInput.pageNumber": str(page_number),
            "itemFilter(0).name": "StartTimeFrom",  # 특정 시간 이후에 올라온 항목
            "itemFilter(0).value": last_update_date,
        }

    response = requests.get(url, headers=headers, params=params)

    if response.status_code == 200:
            data = json.loads(response.text)
            items = data['findItemsByKeywordsResponse'][0]['searchResult'][0].get('item', [])
            
            if not items:
                break
            all_items.extend(items)
            page_number += 1

            # API 호출 제한 체크
            if page_number > 45:
                break
    else:
        print(f"API 호출 실패. 상태 코드: {response.status_code}")
        break
    
product_list = []

# 각 제품 정보 추출
for item in all_items:
    item_id = item['itemId'][0] if 'itemId' in item else "N/A"
    title = item['title'][0] if 'title' in item else "N/A"
    category = item['primaryCategory'][0]['categoryName'][0] if 'primaryCategory' in item else 'N/A'
    price = item['sellingStatus'][0]['currentPrice'][0]['__value__'] if 'sellingStatus' in item else "N/A"
    price = f'$ {price}'

    end_time_str = item['listingInfo'][0]['endTime'][0] if 'listingInfo' in item else "N/A"
    end_time = datetime.strptime(end_time_str, "%Y-%m-%dT%H:%M:%S.%fZ") if end_time_str != "N/A" else "N/A"
    end_time_formatted = end_time.strftime("%Y-%m-%d %H:%M:%S") if end_time != "N/A" else "N/A"  # 원하는 형식으로 변환

    item_url = item['viewItemURL'][0] if 'viewItemURL' in item else "N/A"
    condition = item['condition'][0]['conditionDisplayName'][0] if 'condition' in item else "N/A"
    location = item['location'][0] if 'location' in item else 'N/A'
    shipping = "N/A"
    if 'shippingInfo' in item and len(item['shippingInfo']) > 0:
        if 'shippingServiceCost' in item['shippingInfo'][0]:
            shipping_cost = item['shippingInfo'][0]['shippingServiceCost'][0]['__value__']
            if shipping_cost == "0.0":
                shipping = "Free"
            else:
                shipping = f"${shipping_cost}"
        else:
            shipping = "N/A"
    
    # 제품 정보를 딕셔너리 형태로 추가
    product_list.append({
        "상품 ID": item_id,
        "상품": title,
        "카테고리": category,
        "가격 $": price,
        "경매 마감일": end_time,
        "제품 링크": item_url,
        "상태": condition,
        "판매 국가": location,
        "배송비": shipping
    })

#df = pd.DataFrame(product_list)
#df['Status'] = '새로 추가'

new_df = pd.DataFrame(product_list)
new_df["상품 ID"] = new_df["상품 ID"]

latest_file = find_file.find_latest_file()
latest_file = os.path.join('output', latest_file)
if latest_file:
        existing_df = pd.read_excel(latest_file)
        # 'Newly Added' 상태 제거 (값을 빈 문자열로 변경)
        existing_df['Status'] = existing_df['Status'].replace('새로 추가', '')

current_time = datetime.now()
existing_df = existing_df[existing_df['경매 마감일'] >= current_time.strftime("%Y-%m-%d %H:%M:%S")]

new_products = new_df[~new_df["상품 ID"].isin(existing_df["상품 ID"])]
if not new_products.empty:
    new_products['Status'] = '새로 추가'

updated_df = pd.concat([new_products, existing_df], ignore_index=True)

save_dir = './output'

excel_file = os.path.join(save_dir, f'Scotty Products_{current_date}.xlsx')
updated_df.to_excel(excel_file, index=False)

# openpyxl을 사용하여 엑셀 파일 열기
workbook = load_workbook(excel_file)
sheet = workbook.active

# 각 열의 너비를 자동으로 조정
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter  # 열 이름 (A, B, C, ...)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)  # 여유를 위해 +2 추가
    sheet.column_dimensions[column_letter].width = adjusted_width

url_column_index = 6  # '제품 링크' 열의 인덱스 (A=1, B=2, ..., '제품 링크' 위치에 맞춰 변경)
for row in range(2, sheet.max_row + 1):  # 헤더는 제외하고 2번째 행부터 시작
    cell = sheet.cell(row=row, column=url_column_index)
    url = cell.value  # URL 가져오기
    if url:  # URL이 존재할 경우 하이퍼링크 추가
        cell.hyperlink = url
        cell.value = url
        cell.style = "Hyperlink"

# 엑셀 파일 저장
workbook.save(excel_file)

updated_file = "output/Scotty Products_{current_date}.xlsx"
json_file = "output/update_date.json"

update_data = {
    "updateDate": current_date
}

with open(json_file, 'w') as jsonf:
    json.dump(update_data, jsonf)

# 엑셀 파일 푸시하는 스크립트 예시
subprocess.run(["git", "add", "."])
subprocess.run(["git", "commit", "-m", "Update Excel file and JSON with current date"])
subprocess.run(["git", "push"])

